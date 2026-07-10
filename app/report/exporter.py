"""
报告层 - 选股报告导出
========================================
把软件当前的分析结果一键导出为两种文件,存到项目 reports/ 目录:

  1) Excel(.xlsx)   多 sheet:概览 / 策略排行榜 / 今日推荐 / 选股结果 / 自选持仓
     —— 适合自己做二次筛选、留存、发给别人。
  2) HTML(.html)    自包含单文件,浏览器直接打开,排版美观
     —— 适合快速浏览、打印、分享。涨红跌绿(A股习惯)。

【设计原则】
  · 纯本地生成,不依赖网络,是最稳的功能。
  · 只消费"已经算好"的数据(排行榜/推荐/选股结果/自选),自己不跑策略,
    所以导出很快。数据由调用方(UI)以 snapshot 字典传入。
  · 文件名带时间戳,不会互相覆盖,天然形成历史归档。

对外主入口:
  export_all(snapshot) -> (xlsx_path, html_path)
"""
import os
from datetime import datetime

import pandas as pd

# 报告输出目录:项目根/reports
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
REPORT_DIR = os.path.join(_BASE_DIR, "reports")

# A 股配色(HTML 用)
C_UP = "#d62728"     # 涨/盈利 红
C_DOWN = "#12a05a"   # 跌/亏损 绿
C_FLAT = "#888888"


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _ensure_dir():
    os.makedirs(REPORT_DIR, exist_ok=True)


def build_snapshot(rank_df=None, picks_df=None, results_df=None,
                   results_name="", market=None) -> dict:
    """
    汇总当前分析结果为一个快照 dict,供导出使用。
    调用方(UI)把内存里已算好的结果传进来;自选持仓在这里从 DB 现读。

    参数:
      rank_df      策略排行榜 DataFrame(可空)
      picks_df     今日推荐 DataFrame(可空)
      results_df   当前选股结果 DataFrame(可空)
      results_name 当前选股结果对应的策略名(用于报告标题)
      market       (date, strong: bool) 或 None,大盘状态
    """
    from ..data import database as db

    imap = db.load_industry_map()
    # 自选持仓:现读并算浮动盈亏
    watch_rows = []
    wl = db.load_watchlist()
    if wl is not None and not wl.empty:
        for _, r in wl.iterrows():
            code = r["code"]
            kl = db.load_kline(code)
            cur = float(kl.iloc[-1]["close"]) if kl is not None and not kl.empty else 0.0
            buy = float(r["buy_price"] or 0.0)
            pnl = ((cur - buy) / buy * 100) if buy > 0 else None
            watch_rows.append({
                "code": code, "name": r["name"], "industry": imap.get(code, ""),
                "add_date": r["add_date"], "buy_price": buy, "cur_price": cur,
                "pnl_pct": pnl, "note": r["note"],
            })
    watch_df = pd.DataFrame(watch_rows)

    cache = db.cache_summary()
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "market": market,
        "cache": cache,
        "industry_map": imap,
        "rank_df": rank_df if rank_df is not None else pd.DataFrame(),
        "picks_df": picks_df if picks_df is not None else pd.DataFrame(),
        "results_df": results_df if results_df is not None else pd.DataFrame(),
        "results_name": results_name,
        "watch_df": watch_df,
    }


# ==================== Excel 导出 ====================
def export_excel(snap: dict, path: str = None) -> str:
    """把快照写成多 sheet 的 Excel。返回文件路径。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _ensure_dir()
    path = path or os.path.join(REPORT_DIR, f"选股报告_{_ts()}.xlsx")

    wb = Workbook()
    imap = snap.get("industry_map", {})

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1F4E78")
    up_font = Font(color="C00000")     # 盈利红
    down_font = Font(color="0F8B4C")   # 亏损绿
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    def write_sheet(ws, title, columns, rows, color_rules=None):
        """
        通用写表。
        columns: [(key, 中文列名), ...]
        rows: list[dict]
        color_rules: {列key: 判定函数(value)->'up'/'down'/None}
        """
        color_rules = color_rules or {}
        ws["A1"] = title
        ws["A1"].font = title_font
        ws.append([])
        header = [c[1] for c in columns]
        ws.append(header)
        hr = ws.max_row
        for j, _ in enumerate(columns, start=1):
            cell = ws.cell(row=hr, column=j)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = center
            cell.border = border
        for row in rows:
            vals = [row.get(k, "") for k, _ in columns]
            ws.append(vals)
            rr = ws.max_row
            for j, (k, _) in enumerate(columns, start=1):
                cell = ws.cell(row=rr, column=j)
                cell.border = border
                cell.alignment = center
                rule = color_rules.get(k)
                if rule:
                    tag = rule(row.get(k))
                    if tag == "up":
                        cell.font = up_font
                    elif tag == "down":
                        cell.font = down_font
        # 列宽自适应(粗略)
        for j, (k, name) in enumerate(columns, start=1):
            width = max(len(str(name)) * 2 + 2,
                        *(len(str(r.get(k, ""))) + 2 for r in rows)) if rows \
                else len(str(name)) * 2 + 2
            ws.column_dimensions[get_column_letter(j)].width = min(max(width, 8), 40)

    def sign_up(v):
        try:
            v = float(v)
        except Exception:
            return None
        return "up" if v > 0 else ("down" if v < 0 else None)

    def always_down(v):
        return "down"  # 回撤永远标绿(风险)

    # ---- Sheet 1: 概览 ----
    ws = wb.active
    ws.title = "概览"
    ws["A1"] = "A股选股分析报告"
    ws["A1"].font = Font(bold=True, size=18, color="1F4E78")
    ws["A3"] = "生成时间"
    ws["B3"] = snap["generated_at"]
    m = snap.get("market")
    ws["A4"] = "大盘状态"
    if m and m[0] is not None:
        ws["B4"] = f"{'多头(走强)' if m[1] else '弱势(观望)'}  截至 {m[0]}"
        ws["B4"].font = up_font if m[1] else down_font
    else:
        ws["B4"] = "未知(指数未拉取)"
    c = snap.get("cache", {})
    ws["A5"] = "本地数据"
    ws["B5"] = f"{c.get('codes', 0)} 只 · {c.get('rows', 0)} 行 · 最新 {c.get('last_date', '-')}"
    ws["A6"] = "当前选股策略"
    ws["B6"] = snap.get("results_name", "-") or "-"
    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48

    # ---- Sheet 2: 策略排行榜 ----
    ws = wb.create_sheet("策略排行榜")
    rank = snap["rank_df"]
    if rank is not None and not rank.empty:
        cols = [("rank", "排名"), ("name", "策略"), ("score", "综合分"),
                ("cagr", "年化%"), ("total_return", "总收益%"),
                ("win_rate", "胜率%"), ("max_dd", "回撤%"),
                ("profit_factor", "盈亏比"), ("n_trades", "交易数")]
        write_sheet(ws, "策略排行榜(按综合分)", cols, rank.to_dict("records"),
                    color_rules={"cagr": sign_up, "total_return": sign_up,
                                 "max_dd": always_down})
    else:
        ws["A1"] = "策略排行榜(暂无数据,请先在『今日推荐』生成)"
        ws["A1"].font = title_font

    # ---- Sheet 3: 今日推荐 ----
    ws = wb.create_sheet("今日推荐")
    picks = snap["picks_df"]
    if picks is not None and not picks.empty:
        pk = picks.copy()
        pk["industry"] = pk["code"].map(lambda x: imap.get(x, ""))
        cols = [("code", "代码"), ("name", "名称"), ("industry", "行业"),
                ("close", "现价"), ("strategy", "来源策略"),
                ("strat_rank", "策略排名"), ("hits", "命中"),
                ("score", "得分"), ("reason", "说明")]
        write_sheet(ws, "今日推荐票(最强策略选出)", cols, pk.to_dict("records"))
    else:
        ws["A1"] = "今日推荐(暂无数据,请先在『今日推荐』生成)"
        ws["A1"].font = title_font

    # ---- Sheet 4: 选股结果 ----
    ws = wb.create_sheet("选股结果")
    res = snap["results_df"]
    if res is not None and not res.empty:
        rs = res.copy()
        rs["industry"] = rs["code"].map(lambda x: imap.get(x, ""))
        cols = [("code", "代码"), ("name", "名称"), ("industry", "行业"),
                ("close", "现价"), ("score", "得分"), ("reason", "说明")]
        write_sheet(ws, f"选股结果 · {snap.get('results_name', '')}",
                    cols, rs.to_dict("records"))
    else:
        ws["A1"] = "选股结果(暂无,请先在左侧『开始选股』)"
        ws["A1"].font = title_font

    # ---- Sheet 5: 自选持仓 ----
    ws = wb.create_sheet("自选持仓")
    watch = snap["watch_df"]
    if watch is not None and not watch.empty:
        cols = [("code", "代码"), ("name", "名称"), ("industry", "行业"),
                ("add_date", "加入日"), ("buy_price", "买入价"),
                ("cur_price", "现价"), ("pnl_pct", "浮动盈亏%"), ("note", "备注")]
        write_sheet(ws, "自选持仓(浮动盈亏)", cols, watch.to_dict("records"),
                    color_rules={"pnl_pct": sign_up})
    else:
        ws["A1"] = "自选持仓(暂无)"
        ws["A1"].font = title_font

    wb.save(path)
    return path


# ==================== HTML 导出 ====================
def _fmt_num(v, digits=2):
    try:
        f = float(v)
        return f"{f:.{digits}f}"
    except Exception:
        return "" if v is None else str(v)


def _color_span(v, invert=False, digits=2, suffix=""):
    """根据正负给数字上色(涨红跌绿)。invert=True 时固定绿(用于回撤)。"""
    try:
        f = float(v)
    except Exception:
        return f'<span>{"" if v is None else v}</span>'
    if invert:
        color = C_DOWN
    else:
        color = C_UP if f > 0 else (C_DOWN if f < 0 else C_FLAT)
    sign = "+" if (f > 0 and not invert) else ""
    return f'<span style="color:{color};font-weight:600">{sign}{f:.{digits}f}{suffix}</span>'


def export_html(snap: dict, path: str = None) -> str:
    """把快照写成自包含的 HTML 单文件。返回文件路径。"""
    _ensure_dir()
    path = path or os.path.join(REPORT_DIR, f"选股报告_{_ts()}.html")
    imap = snap.get("industry_map", {})

    m = snap.get("market")
    if m and m[0] is not None:
        mkt_txt = (f'<span style="color:{C_UP}">多头 · 走强</span>' if m[1]
                   else f'<span style="color:{C_DOWN}">弱势 · 观望</span>')
        mkt_txt += f'  <span class="muted">(截至 {m[0]})</span>'
    else:
        mkt_txt = '<span class="muted">未知(指数未拉取)</span>'
    cache = snap.get("cache", {})

    parts = [f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>A股选股报告 {snap['generated_at']}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: "Microsoft YaHei","PingFang SC",sans-serif; margin:0;
         background:#f4f6f9; color:#222; padding:28px; }}
  .wrap {{ max-width:1180px; margin:0 auto; }}
  h1 {{ font-size:24px; margin:0 0 4px; color:#1f4e78; }}
  h2 {{ font-size:17px; margin:30px 0 10px; color:#1f4e78;
        border-left:4px solid #1f4e78; padding-left:9px; }}
  .meta {{ background:#fff; border-radius:10px; padding:16px 20px;
           box-shadow:0 1px 3px rgba(0,0,0,.06); display:flex; flex-wrap:wrap; gap:26px; }}
  .meta div span.k {{ color:#888; font-size:12px; display:block; }}
  .meta div span.v {{ font-size:15px; font-weight:600; }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px;
           overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.06); font-size:13px; }}
  th {{ background:#1f4e78; color:#fff; padding:9px 10px; text-align:center;
        font-weight:600; white-space:nowrap; }}
  td {{ padding:8px 10px; text-align:center; border-bottom:1px solid #eef1f4; }}
  tr:nth-child(even) td {{ background:#fafbfc; }}
  tr:hover td {{ background:#eef5ff; }}
  .muted {{ color:#999; }}
  .tag {{ display:inline-block; background:#eef1f4; border-radius:4px;
          padding:1px 7px; margin:2px; font-size:12px; }}
  .tag.hot {{ background:#fde8e8; color:{C_UP}; font-weight:600; }}
  .empty {{ color:#aaa; background:#fff; padding:16px; border-radius:8px; }}
  .rank1 td {{ background:#fffbe6 !important; font-weight:600; }}
  footer {{ margin-top:34px; color:#aaa; font-size:12px; text-align:center; }}
  .reason {{ text-align:left; color:#555; max-width:340px; }}
</style></head><body><div class="wrap">
<h1>A股选股分析报告</h1>
<div class="meta">
  <div><span class="k">生成时间</span><span class="v">{snap['generated_at']}</span></div>
  <div><span class="k">大盘状态</span><span class="v">{mkt_txt}</span></div>
  <div><span class="k">本地数据</span><span class="v">{cache.get('codes',0)} 只 · 最新 {cache.get('last_date','-')}</span></div>
  <div><span class="k">当前策略</span><span class="v">{snap.get('results_name','-') or '-'}</span></div>
</div>
"""]

    # ---- 策略排行榜 ----
    parts.append("<h2>策略排行榜(按综合表现)</h2>")
    rank = snap["rank_df"]
    if rank is not None and not rank.empty:
        parts.append("<table><tr><th>排名</th><th>策略</th><th>综合分</th>"
                     "<th>年化%</th><th>总收益%</th><th>胜率%</th><th>回撤%</th>"
                     "<th>盈亏比</th><th>交易数</th></tr>")
        for _, r in rank.iterrows():
            cls = ' class="rank1"' if int(r["rank"]) == 1 else ""
            parts.append(
                f"<tr{cls}><td>{int(r['rank'])}</td><td>{r['name']}</td>"
                f"<td>{_fmt_num(r['score'],1)}</td>"
                f"<td>{_color_span(r['cagr'],digits=1)}</td>"
                f"<td>{_color_span(r['total_return'],digits=1)}</td>"
                f"<td>{_fmt_num(r['win_rate'],1)}</td>"
                f"<td>{_color_span(r['max_dd'],invert=True,digits=1)}</td>"
                f"<td>{_fmt_num(r['profit_factor'],2)}</td>"
                f"<td>{int(r['n_trades'])}</td></tr>")
        parts.append("</table>")
    else:
        parts.append('<div class="empty">暂无排行榜数据(请先在软件『今日推荐』页生成)</div>')

    # ---- 热门板块 ----
    picks = snap["picks_df"]
    if picks is not None and not picks.empty:
        from collections import Counter
        cnt = Counter(imap.get(c, "未分类") for c in picks["code"])
        top = cnt.most_common(10)
        parts.append("<h2>热门板块(今日推荐票行业分布)</h2><div>")
        for name, c in top:
            cls = "tag hot" if c >= 2 else "tag"
            parts.append(f'<span class="{cls}">{name} × {c}</span>')
        parts.append("</div>")

    # ---- 今日推荐 ----
    parts.append("<h2>今日推荐票</h2>")
    if picks is not None and not picks.empty:
        parts.append("<table><tr><th>代码</th><th>名称</th><th>行业</th>"
                     "<th>现价</th><th>来源策略</th><th>策略排名</th>"
                     "<th>命中</th><th>得分</th><th>说明</th></tr>")
        for _, r in picks.iterrows():
            ind_txt = imap.get(r["code"], "-")
            hits = int(r["hits"])
            hits_txt = f'<span style="color:{C_UP};font-weight:600">{hits}</span>' if hits >= 2 else str(hits)
            parts.append(
                f"<tr><td>{r['code']}</td><td>{r['name']}</td><td>{ind_txt}</td>"
                f"<td>{_fmt_num(r['close'])}</td><td>{r['strategy']}</td>"
                f"<td>{int(r['strat_rank'])}</td><td>{hits_txt}</td>"
                f"<td>{_fmt_num(r['score'],1)}</td>"
                f"<td class='reason'>{r['reason']}</td></tr>")
        parts.append("</table>")
    else:
        parts.append('<div class="empty">暂无今日推荐(请先在软件『今日推荐』页生成)</div>')

    # ---- 选股结果 ----
    res = snap["results_df"]
    parts.append(f"<h2>选股结果 · {snap.get('results_name','')}</h2>")
    if res is not None and not res.empty:
        parts.append("<table><tr><th>代码</th><th>名称</th><th>行业</th>"
                     "<th>现价</th><th>得分</th><th>说明</th></tr>")
        for _, r in res.iterrows():
            parts.append(
                f"<tr><td>{r['code']}</td><td>{r['name']}</td>"
                f"<td>{imap.get(r['code'],'-')}</td>"
                f"<td>{_fmt_num(r['close'])}</td><td>{_fmt_num(r['score'],1)}</td>"
                f"<td class='reason'>{r['reason']}</td></tr>")
        parts.append("</table>")
    else:
        parts.append('<div class="empty">暂无选股结果(请先在左侧选择策略并『开始选股』)</div>')

    # ---- 自选持仓 ----
    parts.append("<h2>自选持仓</h2>")
    watch = snap["watch_df"]
    if watch is not None and not watch.empty:
        parts.append("<table><tr><th>代码</th><th>名称</th><th>行业</th>"
                     "<th>加入日</th><th>买入价</th><th>现价</th>"
                     "<th>浮动盈亏%</th><th>备注</th></tr>")
        for _, r in watch.iterrows():
            pnl = r["pnl_pct"]
            pnl_txt = _color_span(pnl, digits=2, suffix="%") if pnl is not None \
                else '<span class="muted">观察</span>'
            buy_txt = _fmt_num(r["buy_price"]) if r["buy_price"] > 0 else "-"
            parts.append(
                f"<tr><td>{r['code']}</td><td>{r['name']}</td>"
                f"<td>{r['industry'] or '-'}</td><td>{r['add_date']}</td>"
                f"<td>{buy_txt}</td><td>{_fmt_num(r['cur_price'])}</td>"
                f"<td>{pnl_txt}</td><td class='reason'>{r['note']}</td></tr>")
        parts.append("</table>")
    else:
        parts.append('<div class="empty">暂无自选持仓</div>')

    parts.append('<footer>本报告由 A股选股工具本地生成 · 仅供研究参考,不构成投资建议 · '
                 '涨红跌绿(A股习惯)</footer>')
    parts.append("</div></body></html>")

    with open(path, "w", encoding="utf-8") as f:
        f.write("".join(parts))
    return path


def export_all(snap: dict) -> tuple:
    """一次导出 Excel + HTML,返回 (xlsx_path, html_path)。"""
    xlsx = export_excel(snap)
    html = export_html(snap)
    return xlsx, html
